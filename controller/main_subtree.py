import time
import json
from get_sample import get_sample
from openpyxl import Workbook
from simp_subtree import *
from miasm.expression.simplifications import expr_simp


# modules = [Msynth_simp()]
modules = [Msynth_synth()]
# modules = [Xyntia()]
# modules = [Msynth_synth(), Xyntia(),Msynth_simp()]
# modules = [ Xyntia(),Msynth_synth(),Msynth_simp()]

# modules = [Msynth_synth(), Xyntia(), Msynth_simp()]
# modules_re = [Xyntia(), Msynth_simp()]


def init_excel(ws):
    ws.cell(1,1, "id")
    ws.cell(1,2, "obf_expr")
    ws.cell(1,3, "obf_leng")
    ws.cell(1,4, "obf_var")
    ws.cell(1,5, "result_expr")
    ws.cell(1,6, "result_leng")
    ws.cell(1,7, "result_var")
    ws.cell(1,8, "time")
    ws.cell(1,9, "module")
    ws.cell(1,10, "success")




def write_excel(tmpdict, id, ws):
    ws.cell(id+2, 1, tmpdict["id"])
    for i,k in enumerate(tmpdict):
        if k == "id":
            continue
        ws.cell(id+2,i+2, tmpdict[k])
    ws.cell(id+2,1, tmpdict["id"])
    ws.cell(id+2,2, tmpdict["obf_expr"])
    ws.cell(id+2,3, tmpdict["obf_leng"])
    ws.cell(id+2,4, tmpdict["obf_var"])
    ws.cell(id+2,5, tmpdict["result_expr"])
    ws.cell(id+2,6, tmpdict["result_leng"])
    ws.cell(id+2,7, tmpdict["result_var"])
    ws.cell(id+2,8, tmpdict["time"])
    ws.cell(id+2,9, tmpdict["module"])
    if tmpdict["module"] == "xyntia":
        ws.cell(id+2,10, tmpdict["success"])


def run_PLASynth(samples, outfile,outfile_excel=""):
    file = open(outfile, "w")
    file.write("[\n")
    write_wb = Workbook()
    write_ws = write_wb.active
    # succ_count = 0
    # semantic_count = 0
    subtree = Simp_subtree()

    for i,expr in enumerate(samples):
        if (i!=0):
            file.write(",")
        # if i not in [4,5,6,8,10,11,20,28,29,34,35,41,43,62,63,67,68,69,71,72,74,75,78,81,83,88,96,99,101,102,108,109,112,113,115,116,120,122,125,128,129,132]:
        #     continue
        print(i)
        start_time = time.time()
        subtree.gened = {}
        result = subtree.simplify(expr)
        simplified = ""
        fixed = result
        while simplified != fixed:
            simplified = fixed
            fixed = expr_simp(simplified.replace_expr(subtree.gened))

        # simplified_2, semantic_eq_2 = modules[0].simplify(expr)
        # if semantic_eq_2:
        #     simplified = simplified_2

        endtime = round(time.time() - start_time, 2)

        semantic_eq = modules[0].semantically_equal(expr, simplified)
        # semantic_eq = True
        tmpdict_first = modules[0].getDict_fromExpr(expr, simplified, semantic_eq)
        print(result, subtree.gened)
        print(simplified)
        if tmpdict_first["result_expr"] == "FAIL":
            print("fail")
        tmpdict_first['time'] = endtime
        tmpdict_first["id"] = i
        tmpdict_first["obf_leng"] = expr.length()

        file.write(json.dumps(tmpdict_first, indent=2))

        # write_excel(tmpdict_first,i,write_ws)
    init_excel(write_ws)
    file.write("\n]")
    file.close()



def run(synthesis_module_type, sample_type):
    # filename = "./mbablast_compare/%s_brute.json" % (sample_type)
    # filename = "./mbablast_compare/plasynth_luby_%s_timeout_xyntia5_msynth1360(score10)_simp15_0313.json" % (sample_type)
    filename = "./result/220623[toblast]%s_subexpr_underdepth4.json" % (sample_type)
    # filename = "./result_0124_finaltest/%s_%s_try2.json" % (synthesis_module_type, sample_type)
    run_PLASynth(samples=samples,outfile=filename)

if __name__ == "__main__":
    # pass
    synthesis_module_type = "plasynth" # don't modify
    sample_type = "qsynth" # select sample type {diff, qsynth, tigress, other}

    samples = get_sample(sample_type)
    run(synthesis_module_type, sample_type)
